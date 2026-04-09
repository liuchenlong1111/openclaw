#!/usr/bin/env python3
"""
Extract text content from Excel files (.xlsx, .xls).
"""

import sys
import argparse

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

def extract_excel(file_path):
    """Extract text from Excel file"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"=== Sheet: {sheet_name} ===")
            
            sheet_rows = []
            for row in sheet.iter_rows(values_only=True):
                row_text = []
                for cell in row:
                    if cell is not None:
                        row_text.append(str(cell))
                if row_text:
                    sheet_rows.append(" | ".join(row_text))
            
            text_parts.append("\n".join(sheet_rows))
        
        return "\n\n".join(text_parts)
    except Exception as e:
        print(f"Error extracting Excel: {e}", file=sys.stderr)
        sys.exit(1)

def main():
    parser = argparse.ArgumentParser(description='Extract text from Excel files')
    parser.add_argument('file_path', help='Path to Excel file')
    parser.add_argument('--output', '-o', help='Output file (default: stdout)')
    args = parser.parse_args()
    
    text = extract_excel(args.file_path)
    
    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(text)
        print(f"Extracted text saved to {args.output}", file=sys.stderr)
    else:
        print(text)

if __name__ == '__main__':
    main()
